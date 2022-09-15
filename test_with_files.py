import zipfile
import os
import shutil
import pytest
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import csv

# постусловие - удаление папки, после тестов
@pytest.fixture
def delete_test_folder():
    yield
    shutil.rmtree(os.path.abspath('resources'))

def test_create_and_add_in_zip_file():
    newzip = zipfile.ZipFile(os.path.abspath('file.zip'), "w")
    newzip.write('Тест.pdf')
    newzip.write('Книга1.xlsx')
    newzip.write('Отчет.csv')
    newzip.close()


def test_create_and_move_zip_in_new_folder():
    os.mkdir("resources")
    directory_file = os.path.abspath("file.zip")
    directory_folder = os.path.abspath("resources")
    shutil.move(directory_file, directory_folder)


def test_read_and_check_zip_pdf():
    with zipfile.ZipFile(os.path.abspath("resources/file.zip"), mode='r') as file:
        with file.open("Тест.pdf") as pdf_file:
            pdf_file = PdfReader(pdf_file)
            page = pdf_file.pages[0]
            text = page.extract_text()
            assert "Тест2" in text


def test_read_and_check_in_zip_xlsx():
    with zipfile.ZipFile(os.path.abspath("resources/file.zip"), mode='r') as file:
        with file.open("Книга1.xlsx") as xlsx_file:
            xlsx_file = load_workbook(xlsx_file)
            sheet = xlsx_file.active
            value = sheet.cell(row=3, column=3).value
            assert value == "Сергеев"


def test_read_and_check_in_zip_csv(delete_test_folder):
    with zipfile.ZipFile(os.path.abspath("resources/file.zip")) as file:
        file.extract("Отчет.csv",os.path.abspath("resources"))
        with open("Отчет.csv") as csv_file:
            table = csv.reader(csv_file, delimiter=';')
            for row in table:
                if row[0] == '14.09.2022':
                    break
            row = ', '.join(row)
            assert row == '14.09.2022, 43, 1, 43'
